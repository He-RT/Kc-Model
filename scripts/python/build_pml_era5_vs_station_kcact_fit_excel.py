from pathlib import Path
import numpy as np
import pandas as pd
from scipy import stats

ROOT = Path('/Users/hert/Projects/dcsdxx')
IN = ROOT/'outputs/tables/pml_era5_kcact_vs_station_weather_kcact.csv'
OUT = ROOT/'outputs/tables/pml_era5_vs_station_kcact_fitted_with_r2.xlsx'
OUT.parent.mkdir(parents=True, exist_ok=True)

df = pd.read_csv(IN)
# Ensure date order and stable columns
for c in ['date','date_prev']:
    if c in df.columns:
        df[c] = pd.to_datetime(df[c], errors='coerce')

def fit_stats(d, name):
    x = d['kcact_station_met'].to_numpy(float)
    y = d['kcact_pml_era5'].to_numpy(float)
    n = len(d)
    if n < 3:
        return {
            '拟合范围': name, '样本数n': n, 'Pearson r': np.nan, 'R²': np.nan,
            '斜率 slope': np.nan, '截距 intercept': np.nan,
            '拟合RMSE': np.nan, 'Kcact直接RMSE(y-x)': np.nan,
            'MAE(y-x)': np.nan, 'Bias(y-x)': np.nan,
            '拟合公式': ''
        }
    lr = stats.linregress(x, y)
    pred = lr.slope*x + lr.intercept
    return {
        '拟合范围': name,
        '样本数n': n,
        'Pearson r': lr.rvalue,
        'R²': lr.rvalue**2,
        '斜率 slope': lr.slope,
        '截距 intercept': lr.intercept,
        '拟合RMSE': float(np.sqrt(np.mean((y-pred)**2))),
        'Kcact直接RMSE(y-x)': float(np.sqrt(np.mean((y-x)**2))),
        'MAE(y-x)': float(np.mean(np.abs(y-x))),
        'Bias(y-x)': float(np.mean(y-x)),
        '拟合公式': f"PML/ERA5 Kcact = {lr.slope:.4f} × 站点Kcact + {lr.intercept:.4f}",
    }

# Overall / regional fit
all_s = fit_stats(df, '区域总体/全部站点')
all_slope, all_intercept = all_s['斜率 slope'], all_s['截距 intercept']
df['区域总体拟合Kcact'] = all_slope * df['kcact_station_met'] + all_intercept
# Station-specific fit
station_stats = []
station_params = {}
for st, g in df.groupby('station'):
    s = fit_stats(g, st)
    station_stats.append(s)
    station_params[st] = (s['斜率 slope'], s['截距 intercept'])
df['分站点拟合Kcact'] = [station_params[r.station][0] * r.kcact_station_met + station_params[r.station][1] for r in df.itertuples()]
df['区域拟合残差'] = df['kcact_pml_era5'] - df['区域总体拟合Kcact']
df['分站点拟合残差'] = df['kcact_pml_era5'] - df['分站点拟合Kcact']
df['Kcact差值(PML区域-站点)'] = df['kcact_pml_era5'] - df['kcact_station_met']

stats_df = pd.DataFrame([all_s] + station_stats)

# Human-readable data table
rename = {
    'station': '站点', 'date': '日期', 'date_prev': '窗口开始日期',
    'station_eta_mm_d': '站点实测ETa(mm/d)',
    'met_station_et0_mm_d': '气象站ET0(mm/d)',
    'pml_eta_mm_d': 'PML ETa(mm/d)',
    'era5_et0_mm_d': 'ERA5 ET0(mm/d)',
    'kcact_station_met': '站点Kcact=实测ETa/气象站ET0',
    'kcact_pml_era5': '区域Kcact=PML ETa/ERA5 ET0',
    'window_days': '窗口天数',
    'n_pml_overlap': 'PML重叠影像数',
    'pml_overlap_days': 'PML重叠天数',
}
cols = [
    'station','date','date_prev','window_days',
    'station_eta_mm_d','met_station_et0_mm_d','pml_eta_mm_d','era5_et0_mm_d',
    'kcact_station_met','kcact_pml_era5',
    '区域总体拟合Kcact','区域拟合残差','分站点拟合Kcact','分站点拟合残差','Kcact差值(PML区域-站点)',
    'n_pml_overlap','pml_overlap_days'
]
data_out = df[[c for c in cols if c in df.columns]].rename(columns=rename).sort_values(['站点','日期'])

summary = pd.DataFrame({
    '项目': [
        'x轴/站点Kcact', 'y轴/区域Kcact', '区域总体拟合Kcact', '分站点拟合Kcact',
        '样本筛选', '用途'
    ],
    '说明': [
        '站点实测 ETa / 气象站 ET0',
        'PML-V2.2a ETa / ERA5 ET0',
        '用全部站点样本统一拟合得到的 y_hat，表示区域尺度总体关系',
        '每个站点单独拟合得到的 y_hat，表示站点内关系',
        '同一站点、同一8日观测窗口对齐；剔除缺失、非8日窗口、ET0<=0.1和极端Kcact',
        '可用于报告中说明站点尺度与区域遥感尺度 Kcact 的一致性及偏差'
    ]
})


from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(OUT, engine='openpyxl', datetime_format='yyyy-mm-dd') as writer:
    summary.to_excel(writer, sheet_name='说明', index=False)
    stats_df.to_excel(writer, sheet_name='R方与拟合统计', index=False)
    data_out.to_excel(writer, sheet_name='拟合后Kcact明细', index=False)

wb = load_workbook(OUT)
header_fill = PatternFill('solid', fgColor='D9EAF7')
thin = Side(style='thin', color='BFBFBF')
header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color='1F4E79')
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # widths
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:200])
        width = min(max(max_len * 1.15 + 2, 10), 35)
        if ws.title == '说明' and letter == 'B': width = 85
        ws.column_dimensions[letter].width = width
    if ws.title == '说明':
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    if ws.title == '拟合后Kcact明细':
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (float, int)) and cell.column > 3:
                    cell.number_format = '0.000'
                if cell.column in [2,3]:
                    cell.number_format = 'yyyy-mm-dd'
    if ws.title == 'R方与拟合统计':
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '0.000'

# Add scatter chart on stats sheet
ws_stats = wb['R方与拟合统计']
ws_data = wb['拟合后Kcact明细']
headers = [c.value for c in ws_data[1]]
x_col = headers.index('站点Kcact=实测ETa/气象站ET0') + 1
y_col = headers.index('区域Kcact=PML ETa/ERA5 ET0') + 1
fit_col = headers.index('区域总体拟合Kcact') + 1
n_rows = ws_data.max_row
chart = ScatterChart()
chart.title = '区域Kcact 与站点Kcact 拟合'
chart.x_axis.title = '站点Kcact = 实测ETa / 气象站ET0'
chart.y_axis.title = '区域Kcact = PML ETa / ERA5 ET0'
chart.width = 18
chart.height = 11
xvalues = Reference(ws_data, min_col=x_col, min_row=2, max_row=n_rows)
yvalues = Reference(ws_data, min_col=y_col, min_row=2, max_row=n_rows)
ser = Series(yvalues, xvalues, title='原始Kcact点')
ser.graphicalProperties.line.noFill = True
ser.marker.symbol = 'circle'
ser.marker.size = 4
ser.marker.graphicalProperties.solidFill = '5DA5DA'
ser.marker.graphicalProperties.line.solidFill = 'FFFFFF'
chart.series.append(ser)
fit_values = Reference(ws_data, min_col=fit_col, min_row=2, max_row=n_rows)
ser2 = Series(fit_values, xvalues, title=f"区域总体拟合 R²={all_s['R²']:.3f}")
ser2.graphicalProperties.line.solidFill = '1F4E79'
ser2.graphicalProperties.line.width = 22000
ser2.marker.symbol = 'none'
chart.series.append(ser2)
ws_stats.add_chart(chart, 'M2')

# Add title note at top of stats? leave table clean.
wb.save(OUT)
print(OUT)
print(stats_df.to_string(index=False))
